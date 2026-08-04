"""v1.44 §3.4 — "Preview in Excel" (pre-approval) + trailer Active/Inactive flag.

Journey 1 (preview): /mes-app/costings/new → the embed header shows the new
"Preview in Excel" button LEFT of Reload → clicking BEFORE any calculation
toasts "Calculate first" inside the calculator iframe → select the staged body
type + nudge a dim (live-calc) → click again → a real download lands, named
"Testing - {body} - {yyyy-mm-dd}.xlsx", whose sheet is headed
"Testing — {body}" with NO Report#/quote/customer — and the calculations table
gained no row (the preview consumes nothing).

Journey 2 (active flag): Admin / Trailer Templates → select the staged body →
"Set Inactive" (confirm modal) → badge appears; a FRESH /calculator load omits
it from the Body Type dropdown; a ?edit= reopen of a pending costing on the
now-inactive body STILL WORKS (fallback option labelled "(inactive)"); "Set
Active" again → the dropdown has it back.

Selector policy: calculator + admin pages are Jinja (element IDs); the embed
header button is queried by role+name. No wait_for_function anywhere (CSP has
no unsafe-eval). Marker J144XP; purge at setup AND teardown.
"""
from __future__ import annotations

import datetime as _dt
import io
import json
import re

import pytest
from playwright.sync_api import Page, expect

from _common import admin_session, shot  # noqa: E402  (sys.path set in conftest)

T = 15_000
JOURNEY = "excel_preview_active_flag"
MARK = "J144XP"


def _purge(db) -> None:
    from sqlalchemy import text
    db.execute(text(
        "DELETE FROM icb_costings.calculations WHERE trailer_type_id IN "
        "(SELECT id FROM icb_costings.trailer_types WHERE name LIKE :m)"), {"m": f"{MARK}%"})
    db.execute(text(
        "DELETE FROM icb_costings.bill_of_materials WHERE material_id IN "
        "(SELECT id FROM icb_costings.materials WHERE name LIKE :m)"), {"m": f"{MARK}%"})
    db.execute(text("DELETE FROM icb_costings.materials WHERE name LIKE :m"), {"m": f"{MARK}%"})
    db.execute(text("DELETE FROM icb_costings.trailer_types WHERE name LIKE :m"), {"m": f"{MARK}%"})
    db.commit()


@pytest.fixture(scope="module")
def staged():
    """Calculable marker trailer: one plain always-include BOM row (price 250,
    formula '1') + a pending costing record for the ?edit= reopen check."""
    from app.database import BillOfMaterial, CalculationRecord, Material, SessionLocal, TrailerType, User
    with SessionLocal() as db:
        _purge(db)
        trailer = TrailerType(name=f"{MARK} PREVIEW BODY", is_active=True,
                              default_length=6.0, default_width=2.4, default_height=2.4)
        db.add(trailer)
        db.flush()
        mat = Material(name=f"{MARK} PANEL SHEET", unit_of_measure="m2", price_per_unit=250.0)
        db.add(mat)
        db.flush()
        db.add(BillOfMaterial(trailer_type_id=trailer.id, material_id=mat.id,
                              formula_expression="1", waste_percentage=0,
                              bom_section="PANELS", sort_order=1))
        admin = db.query(User).filter_by(username="admin").first()
        rec = CalculationRecord(
            trailer_type_id=trailer.id, user_id=admin.id, status="pending",
            dimensions_json=json.dumps({"length": 6.0, "width": 2.4, "height": 2.4}),
            result_json=json.dumps({
                "items": [{"category": "PANELS", "material": f"{MARK} PANEL SHEET",
                           "material_code": "", "formula": "1", "quantity": 1.0,
                           "unit": "m2", "unit_price": 250.0, "waste_pct": 0,
                           "line_cost": 250.0, "last_updated": None}],
                "category_totals": {"PANELS": 250.0},
                "grand_total": 250.0, "cost_per_sqm": 17.36, "selling_price": 455.0,
            }))
        db.add(rec)
        db.commit()
        ids = {"trailer": trailer.id, "rec": rec.id, "name": trailer.name}
    yield ids
    from app.database import SessionLocal as SL
    with SL() as db:
        _purge(db)


def _record_count() -> int:
    from app.database import CalculationRecord, SessionLocal
    with SessionLocal() as db:
        return db.query(CalculationRecord).count()


def _click_until_toast(page: Page, button, frame, text: str, attempts: int = 6) -> None:
    """postMessage is one-shot: if the iframe's listener isn't registered yet the
    click is silently lost. Re-click (bounded) until the calculator toasts."""
    toast = frame.locator("#toast-container .toast-msg", has_text=text)
    for _ in range(attempts):
        button.click()
        try:
            expect(toast.first).to_be_visible(timeout=2_500)
            return
        except AssertionError:
            continue
    raise AssertionError(f"toast {text!r} never appeared after {attempts} clicks")


def test_preview_in_excel_download(page: Page, live_server: str, staged, tmp_path) -> None:
    ids = staged
    admin_session(page, base=live_server)   # base= matters under MES_BASE (banked)
    page.goto("/mes-app/costings/new")

    btn = page.get_by_role("button", name="Preview in Excel")
    expect(btn).to_be_visible(timeout=T)
    frame = page.frame_locator("iframe[title='Calculator (live costing app)']")
    expect(frame.locator("#trailer-select")).to_be_visible(timeout=30_000)
    shot(page, "01-button-in-header", journey=JOURNEY)

    # ── Nothing calculated yet → "Calculate first", no download ──────────────
    _click_until_toast(page, btn, frame, "Calculate first")
    shot(page, "02-calculate-first-toast", journey=JOURNEY)

    # ── Calculate: pick the staged body; a dim nudge fires the live recalc ───
    frame.locator("#trailer-select").select_option(str(ids["trailer"]))
    frame.locator("#f-length").fill("6.5")
    expect(frame.locator("#approve-btn")).to_be_enabled(timeout=30_000)

    before = _record_count()
    with page.expect_download(timeout=T) as dl_info:
        btn.click()
    download = dl_info.value
    today = _dt.date.today().strftime("%Y-%m-%d")
    assert download.suggested_filename == f"Testing - {ids['name']} - {today}.xlsx", \
        download.suggested_filename

    target = tmp_path / download.suggested_filename
    download.save_as(target)
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(target.read_bytes()))
    ws = wb.active
    assert ws["A1"].value == f"Testing — {ids['name']}"
    flat = "|".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    assert "Report #" not in flat and "Client:" not in flat
    assert f"{MARK} PANEL SHEET" in flat

    assert _record_count() == before, "preview must write NOTHING to the DB"
    shot(page, "03-downloaded", journey=JOURNEY)


def test_active_flag_toggle_and_inactive_edit_reopen(page: Page, live_server: str, staged) -> None:
    ids = staged
    admin_session(page, base=live_server)   # base= matters under MES_BASE (banked)

    # ── Admin / Trailer Templates: Set Inactive (with confirm) ───────────────
    page.goto("/admin/templates")
    row = page.locator(f"#tt-{ids['trailer']}")
    expect(row).to_be_visible(timeout=T)
    row.click()
    toggle = page.locator("#btn-toggle-active")
    expect(toggle).to_have_text("Set Inactive", timeout=T)
    toggle.click()
    expect(page.locator("#modal-confirm")).to_be_visible(timeout=T)
    page.locator("#confirm-ok").click()
    expect(page.locator(f"#tt-{ids['trailer']}")).to_contain_text("Inactive", timeout=T)
    expect(toggle).to_have_text("Set Active")
    shot(page, "04-set-inactive-badge", journey=JOURNEY)

    # ── Fresh calculator load: the body type is GONE from the dropdown ───────
    page.goto("/calculator")
    expect(page.locator("#trailer-select")).to_be_visible(timeout=T)
    expect(page.locator(f"#trailer-select option[value='{ids['trailer']}']")).to_have_count(0)
    shot(page, "05-dropdown-omits-inactive", journey=JOURNEY)

    # ── ?edit= reopen of a pending costing on the INACTIVE body still works ──
    page.goto(f"/calculator?edit={ids['rec']}")
    opt = page.locator(f"#trailer-select option[value='{ids['trailer']}']")
    expect(opt).to_have_count(1, timeout=30_000)
    expect(opt).to_have_text(f"{ids['name']} (inactive)")
    expect(page.locator("#trailer-select")).to_have_value(str(ids["trailer"]))
    expect(page.locator("#edit-mode-banner")).to_be_visible(timeout=30_000)
    shot(page, "06-inactive-edit-reopen", journey=JOURNEY)

    # ── Toggle back Active → dropdown has it again ───────────────────────────
    page.goto("/admin/templates")
    row = page.locator(f"#tt-{ids['trailer']}")
    expect(row).to_be_visible(timeout=T)
    row.click()
    toggle = page.locator("#btn-toggle-active")
    expect(toggle).to_have_text("Set Active", timeout=T)
    toggle.click()
    expect(page.locator("#modal-confirm")).to_be_visible(timeout=T)
    page.locator("#confirm-ok").click()
    expect(toggle).to_have_text("Set Inactive", timeout=T)

    page.goto("/calculator")
    expect(page.locator("#trailer-select")).to_be_visible(timeout=T)
    expect(page.locator(f"#trailer-select option[value='{ids['trailer']}']")).to_have_count(1)
    shot(page, "07-active-again", journey=JOURNEY)
