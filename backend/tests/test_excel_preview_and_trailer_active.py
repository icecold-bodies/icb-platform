"""v1.44 — "Preview in Excel" (pre-approval) + Trailer template Active/Inactive flag.

Covers:
  1. POST /api/export/excel-preview — valid xlsx, "Testing — {body type}" heading,
     NO report/quote number, NO customer block, excluded rows stripped, nothing
     written to the DB, and the 400 / 401 / 403 guards.
  2. GET /results/{id}/export/excel — REGRESSION-LOCKED: the context-dict builder
     refactor must keep the saved-record export byte-identical. The lock is a
     normalized hash: sha256 over the zip entries' *decompressed* contents in
     sorted name order, excluding docProps/core.xml (openpyxl stamps save-time
     created/modified timestamps there), so it is stable across runs, platforms
     and zlib builds. The constant below was captured from the post-refactor
     builder, which was itself proven byte-identical to the PRE-refactor route
     on live dev records (1642 + 1639, both highlight modes — see the PR body).
  3. /api/trailers Active/Inactive surface: default list stays active-only (the
     calculators' dropdown source); include_inactive=1 is admin-gated and hides
     soft-deleted "[deleted-…]" rows; GET /api/trailers/{id} serves inactive rows
     (the calculator's ?edit= reopen fallback); PUT toggles is_active; and the
     name-clash checks now cover Inactive templates.

Sessions are real UserSession rows sent via a raw Cookie header — the routes
under test use inline get_current_user / require_admin chokepoints, which
dependency_overrides never reach (banked pattern).
"""
import hashlib
import io
import json
import uuid
import zipfile
from datetime import datetime

import pytest

# ── deterministic fixture inputs (fixed values → fixed regression hash) ───────
FIXED_RECORD_ID = 900144            # explicit PK: the subtitle embeds "Report #{id}"
FIXED_CREATED_AT = datetime(2026, 1, 15, 10, 30, 0)
FIXED_TT_NAME = "V144 REGRESSION BODY"

RESULT_FIXTURE = {
    "items": [
        {"category": "FLOOR", "material": "TEST FLOOR SHEET", "material_code": "TFS-1",
         "formula": "L*W", "quantity": 12.5, "unit": "m2", "unit_price": 100.0,
         "waste_pct": 5, "line_cost": 1312.5, "last_updated": None},
        {"category": "FLOOR", "material": "TEST FLOOR GLUE", "material_code": "TFG-1",
         "formula": "L*W*0.2", "quantity": 2.5, "unit": "kg", "unit_price": 50.0,
         "waste_pct": 0, "line_cost": 125.0, "last_updated": None},
        {"category": "OPTIONAL EXTRAS", "material": "TEST EXTRA", "material_code": "",
         "formula": "1", "quantity": 1.0, "unit": "ea", "unit_price": 400.0,
         "waste_pct": 0, "line_cost": 400.0, "section_is_optional": True,
         "last_updated": None},
        # excluded row: must be stripped from BOTH the saved export and the preview
        {"category": "OPTIONAL EXTRAS", "material": "EXCLUDED EXTRA", "material_code": "",
         "formula": "1", "quantity": 1.0, "unit": "ea", "unit_price": 999.0,
         "waste_pct": 0, "line_cost": 999.0, "section_is_optional": True,
         "excluded": True, "last_updated": None},
    ],
    "category_totals": {"FLOOR": 1437.5, "OPTIONAL EXTRAS": 400.0},
    "cost_per_sqm": 114.99,
    "grand_total": 1837.5,
    "profit_margin": 20,
    "profit_amount": 367.5,
    "ratio_amount": 1002.27,
    "ratio_label": "55%",
    "selling_price": 3341.0,
    "chassis": {"items": [{"kind": "Axle", "label": "TEST AXLE", "qty": 2,
                           "unit_price": 500.0, "line_cost": 1000.0}],
                "subtotal": 1000.0, "axle_count": 2, "tyre_style": "single",
                "tyre_count": 4, "length": 7.5},
}
DIMS_FIXTURE = {"length": 7.5, "width": 2.5, "height": 2.6, "num_axles": 2,
                "num_doors": 2, "insulation_thickness": 0.1}

# Captured from the deterministic record above (see module docstring for why
# this is trustworthy). If openpyxl is ever bumped from ==3.1.5 this constant
# may legitimately change — re-capture it in the same PR as the bump.
REGRESSION_NHASH = "8c4b307812bc9102c9c6e2d0e5cae1b43124b15ebed6aeaa4e593bc0e4a780a8"


def normalized_xlsx_hash(xlsx_bytes: bytes) -> str:
    zf = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
    h = hashlib.sha256()
    for name in sorted(zf.namelist()):
        if name == "docProps/core.xml":
            continue
        h.update(name.encode())
        h.update(b"\x00")
        h.update(zf.read(name))
    return h.hexdigest()


def _sheet_flat_text(xlsx_bytes: bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    flat = "|".join(str(c.value) for row in ws.iter_rows() for c in row
                    if c.value is not None)
    return ws, flat


# ── fixtures ──────────────────────────────────────────────────────────────────
@pytest.fixture(scope="module")
def app_mod():
    import app.main as m
    from starlette.testclient import TestClient
    with TestClient(m.app) as _c:   # triggers startup → seeds admin + permissions
        yield m


@pytest.fixture(scope="module")
def client(app_mod):
    from starlette.testclient import TestClient
    with TestClient(app_mod.app) as c:
        yield c


def _make_session(username: str) -> dict:
    """Real UserSession + raw Cookie header (+ matching CSRF token for POSTs)."""
    from app.database import SessionLocal, User, UserSession
    sid = f"v144-{uuid.uuid4().hex[:12]}"
    csrf = f"csrf-{sid}"
    with SessionLocal() as db:
        u = db.query(User).filter_by(username=username).first()
        assert u, f"user {username!r} missing"
        db.merge(UserSession(id=sid, user_id=u.id, role=u.role,
                             expires_at=None, csrf_token=csrf))
        db.commit()
    return {"Cookie": f"session_id={sid}", "X-CSRF-Token": csrf}


@pytest.fixture(scope="module")
def admin_headers(app_mod):
    return _make_session("admin")


@pytest.fixture(scope="module")
def planner_headers(app_mod):
    """Throwaway planner: no export.excel ({admin, full}) and no admin rights."""
    from app.database import SessionLocal, User
    uname = f"t_planner_{uuid.uuid4().hex[:6]}"
    with SessionLocal() as db:
        db.add(User(username=uname, password_hash="x", role="planner"))
        db.commit()
    headers = _make_session(uname)
    yield headers
    from app.database import UserSession
    with SessionLocal() as db:
        db.query(UserSession).filter_by(
            id=headers["Cookie"].split("session_id=")[1]).delete()
        db.query(User).filter_by(username=uname).delete()
        db.commit()


@pytest.fixture(scope="module")
def seeded(app_mod):
    """Deterministic trailer + calculation record (fixed PK, name, timestamp)."""
    from app.database import SessionLocal, TrailerType, CalculationRecord, User
    with SessionLocal() as db:
        admin = db.query(User).filter_by(username="admin").first()
        tt = db.query(TrailerType).filter_by(name=FIXED_TT_NAME).first()
        if not tt:
            tt = TrailerType(name=FIXED_TT_NAME, description="v1.44 regression fixture")
            db.add(tt)
            db.flush()
        rec = db.query(CalculationRecord).filter_by(id=FIXED_RECORD_ID).first()
        if not rec:
            rec = CalculationRecord(
                id=FIXED_RECORD_ID,
                trailer_type_id=tt.id, user_id=admin.id,
                dimensions_json=json.dumps(DIMS_FIXTURE),
                result_json=json.dumps(RESULT_FIXTURE),
                created_at=FIXED_CREATED_AT, status="pending", is_repair=False)
            db.add(rec)
        db.commit()
        ids = {"tt_id": tt.id, "rec_id": FIXED_RECORD_ID}
    yield ids
    with SessionLocal() as db:
        db.query(CalculationRecord).filter_by(id=FIXED_RECORD_ID).delete()
        db.query(TrailerType).filter_by(id=ids["tt_id"]).delete()
        db.commit()


def _record_count() -> int:
    from app.database import SessionLocal, CalculationRecord
    with SessionLocal() as db:
        return db.query(CalculationRecord).count()


# ── 1. saved-record export regression lock ────────────────────────────────────
def test_saved_export_regression_hash(client, admin_headers, seeded):
    r = client.get(f"/results/{seeded['rec_id']}/export/excel", headers=admin_headers)
    assert r.status_code == 200, r.text
    ws, flat = _sheet_flat_text(r.content)
    assert ws["A1"].value == "TRAILER MANUFACTURING COST REPORT"
    assert f"Report #{seeded['rec_id']}" in (ws["A2"].value or "")
    assert "15 January 2026" in (ws["A2"].value or "")
    assert "EXCLUDED EXTRA" not in flat          # strip_excluded_items on saved path
    assert normalized_xlsx_hash(r.content) == REGRESSION_NHASH, (
        "saved-record Excel export changed byte-wise — if this is deliberate, "
        "re-capture REGRESSION_NHASH in the same change")


# ── 2. excel-preview endpoint ─────────────────────────────────────────────────
def test_preview_returns_testing_workbook(client, admin_headers, seeded):
    r = client.post("/api/export/excel-preview", headers=admin_headers,
                    json={"result": RESULT_FIXTURE, "dims": DIMS_FIXTURE,
                          "trailer_type_id": seeded["tt_id"],
                          "trailer_name": "CLIENT LABEL MUST LOSE"})
    assert r.status_code == 200, r.text
    disp = r.headers.get("content-disposition", "")
    assert f"Testing - {FIXED_TT_NAME} - " in disp and disp.endswith('.xlsx"')
    ws, flat = _sheet_flat_text(r.content)
    assert ws["A1"].value == f"Testing — {FIXED_TT_NAME}"   # DB name wins
    assert "Report #" not in flat and "Client:" not in flat
    assert ws["A3"].value in (None, "")
    assert "TEST FLOOR SHEET" in flat and "TEST AXLE" in flat
    assert "EXCLUDED EXTRA" not in flat          # strip_excluded_items on preview too
    # Preview spec block (Michael 4 Aug): L/W/H only — no axles/doors/insulation grid.
    assert ws["A4"].value == "DIMENSIONS & BODY OPTIONS"
    assert "Num Axles" not in flat and "Num Doors" not in flat
    assert "Insulation Thickness (m)" not in flat


@pytest.fixture(scope="module")
def optioned(app_mod):
    """Trailer with body-option pairs + floor type + plain rows whose
    sort_orders REVERSE the result-item order — proves the spec block and the
    on-screen (sheet) ordering + literal section totals."""
    from app.database import BillOfMaterial, Material, SessionLocal, TrailerType
    name = f"V144 OPTIONED BODY {uuid.uuid4().hex[:6]}"
    with SessionLocal() as db:
        tt = TrailerType(name=name)
        db.add(tt)
        db.flush()

        def _mat(mname):
            m = Material(name=mname, unit_of_measure="ea", price_per_unit=10.0)
            db.add(m)
            db.flush()
            return m

        def _row(mname, **kw):
            r = BillOfMaterial(trailer_type_id=tt.id, material_id=_mat(mname).id,
                               formula_expression="1", waste_percentage=0, **kw)
            db.add(r)
            db.flush()
            return r

        front_eps = _row(f"{name} FRONT EPS", is_body_option=True,
                         body_option_group="FRONT", body_option_subgroup="INSULATION",
                         variable_value=0.06)
        front_pu = _row(f"{name} FRONT PU", is_body_option=True,
                        body_option_group="FRONT", body_option_subgroup="INSULATION",
                        variable_value=0.0)
        drd_eps = _row(f"{name} DRD EPS", is_body_option=True,
                       body_option_group="DRD", body_option_subgroup="INSULATION",
                       variable_value=0.05)
        drd_pu = _row(f"{name} DRD PU", is_body_option=True,
                      body_option_group="DRD", body_option_subgroup="INSULATION",
                      variable_value=0.0)
        rice = _row(f"{name} RICE GRAIN FLOOR", is_body_option=True,
                    body_option_group="FLOOR")
        # Plain rows: sheet order says ZLATE (sort 1) before AEARLY (sort 2) —
        # the OPPOSITE of both alphabetical and the result-item order below.
        zlate = _row(f"{name} ZLATE MAT", bom_section="ZLATE", sort_order=1)
        aearly = _row(f"{name} AEARLY MAT", bom_section="AEARLY", sort_order=2)
        db.commit()
        ids = {"tt_id": tt.id, "name": name,
               "front_eps": front_eps.id, "drd_eps": drd_eps.id, "rice": rice.id,
               "zlate": zlate.id, "aearly": aearly.id}
    yield ids
    with SessionLocal() as db:
        from sqlalchemy import text
        db.execute(text(
            "DELETE FROM icb_costings.bill_of_materials WHERE trailer_type_id = :t"),
            {"t": ids["tt_id"]})
        db.execute(text(
            "DELETE FROM icb_costings.materials WHERE name LIKE :m"), {"m": f"{name}%"})
        db.execute(text(
            "DELETE FROM icb_costings.trailer_types WHERE id = :t"), {"t": ids["tt_id"]})
        db.commit()


def test_preview_spec_block_totals_and_screen_order(client, admin_headers, optioned):
    """Michael 4 Aug adjustments: (1) spec block = L/W/H + selected body options,
    (2) literal numeric section totals on header rows, (3) items in the costings
    page's sheet order (BOM sort_order, not result order)."""
    n = optioned["name"]
    result = {
        # Result order = AEARLY first; sheet order must flip it to ZLATE first.
        "items": [
            {"category": "AEARLY", "material": f"{n} AEARLY MAT", "material_code": "",
             "formula": "1", "quantity": 1.0, "unit": "ea", "unit_price": 10.0,
             "waste_pct": 0, "line_cost": 111.0, "bom_id": optioned["aearly"],
             "last_updated": None},
            {"category": "ZLATE", "material": f"{n} ZLATE MAT", "material_code": "",
             "formula": "1", "quantity": 1.0, "unit": "ea", "unit_price": 10.0,
             "waste_pct": 0, "line_cost": 222.5, "bom_id": optioned["zlate"],
             "last_updated": None},
        ],
        "category_totals": {"AEARLY": 111.0, "ZLATE": 222.5},
        "grand_total": 333.5,
        "body_variables": {f"{n} FRONT EPS": 0.06, f"{n} FRONT PU": 0.0,
                           f"{n} DRD EPS": 0.05, f"{n} DRD PU": 0.0},
    }
    sel = {str(optioned["front_eps"]): True, str(optioned["drd_eps"]): True,
           str(optioned["rice"]): True}
    r = client.post("/api/export/excel-preview", headers=admin_headers,
                    json={"result": result, "dims": {"length": 6.0, "width": 2.4,
                                                     "height": 2.4, "num_axles": 3},
                          "trailer_type_id": optioned["tt_id"],
                          "body_option_selections": sel,
                          "drd_srd": {"DRD": True},
                          "bom_sort_mode": "sheet"})
    assert r.status_code == 200, r.text
    ws, flat = _sheet_flat_text(r.content)

    # (1) spec block: L/W/H + the selected options; no axles even though sent.
    assert ws["A4"].value == "DIMENSIONS & BODY OPTIONS"
    assert ws["A5"].value == "Length (m)" and ws["E5"].value == "Height (m)"
    assert "Num Axles" not in flat
    spec_pairs = {ws.cell(row=r_, column=1).value: ws.cell(row=r_, column=2).value
                  for r_ in range(6, 10)}
    assert spec_pairs.get("DOOR TYPE") == "DRD — EPS (0.050 m)"
    assert spec_pairs.get("FRONT") == "EPS (0.060 m)"
    assert spec_pairs.get("FLOOR TYPE") == f"{n} RICE GRAIN FLOOR"

    # (3) sheet order: ZLATE's section header row precedes AEARLY's.
    col_a = [ws.cell(row=r_, column=1).value or "" for r_ in range(1, ws.max_row + 1)]
    assert col_a.index("ZLATE") < col_a.index("AEARLY")

    # (2) literal numeric totals on the section header rows (no =SUM text).
    z_hdr = col_a.index("ZLATE") + 1
    a_hdr = col_a.index("AEARLY") + 1
    assert ws.cell(row=z_hdr, column=9).value == 222.5
    assert ws.cell(row=a_hdr, column=9).value == 111.0
    assert "=SUM" not in flat


def test_preview_writes_nothing(client, admin_headers, seeded):
    before = _record_count()
    r = client.post("/api/export/excel-preview", headers=admin_headers,
                    json={"result": RESULT_FIXTURE, "dims": DIMS_FIXTURE,
                          "trailer_type_id": seeded["tt_id"]})
    assert r.status_code == 200
    assert _record_count() == before


def test_preview_name_fallback_without_tt(client, admin_headers):
    r = client.post("/api/export/excel-preview", headers=admin_headers,
                    json={"result": RESULT_FIXTURE, "dims": {},
                          "trailer_name": "FALLBACK LABEL"})
    assert r.status_code == 200
    ws, _ = _sheet_flat_text(r.content)
    assert ws["A1"].value == "Testing — FALLBACK LABEL"


def test_preview_guards(client, admin_headers, planner_headers):
    r = client.post("/api/export/excel-preview", headers=admin_headers,
                    json={"result": None})
    assert r.status_code == 400
    r = client.post("/api/export/excel-preview", headers=admin_headers,
                    json={"result": {"items": []}})
    assert r.status_code == 400
    r = client.post("/api/export/excel-preview", json={"result": RESULT_FIXTURE})
    assert r.status_code == 401                  # no session
    r = client.post("/api/export/excel-preview", headers=planner_headers,
                    json={"result": RESULT_FIXTURE})
    assert r.status_code == 403                  # planner lacks export.excel


# ── 3. trailer Active/Inactive surface ────────────────────────────────────────
@pytest.fixture()
def flag_trailer(client, admin_headers):
    """Throwaway trailer created via the API; hard-deleted afterwards."""
    name = f"V144 FLAG BODY {uuid.uuid4().hex[:6]}"
    r = client.post("/api/trailers", headers=admin_headers,
                    json={"name": name, "description": "v1.44 flag test"})
    assert r.status_code == 200, r.text
    tid = r.json()["id"]
    yield {"id": tid, "name": name}
    from app.database import SessionLocal, TrailerType
    with SessionLocal() as db:
        db.query(TrailerType).filter_by(id=tid).delete()
        db.commit()


def _trailer_ids(client, headers, qs=""):
    r = client.get(f"/api/trailers{qs}", headers=headers)
    assert r.status_code == 200, r.text
    return {t["id"]: t for t in r.json()}


def test_toggle_hides_from_default_list_only(client, admin_headers, flag_trailer):
    tid = flag_trailer["id"]
    assert tid in _trailer_ids(client, admin_headers)
    r = client.put(f"/api/trailers/{tid}", headers=admin_headers,
                   json={"is_active": False})
    assert r.status_code == 200
    assert tid not in _trailer_ids(client, admin_headers)            # dropdown source
    everything = _trailer_ids(client, admin_headers, "?include_inactive=1")
    assert tid in everything and everything[tid]["is_active"] is False
    single = client.get(f"/api/trailers/{tid}", headers=admin_headers)
    assert single.status_code == 200 and single.json()["is_active"] is False
    r = client.put(f"/api/trailers/{tid}", headers=admin_headers,
                   json={"is_active": True})
    assert r.status_code == 200
    assert tid in _trailer_ids(client, admin_headers)                # back again


def test_include_inactive_is_admin_gated(client, planner_headers):
    r = client.get("/api/trailers?include_inactive=1", headers=planner_headers)
    assert r.status_code in (401, 403)
    r = client.get("/api/trailers", headers=planner_headers)         # default: fine
    assert r.status_code == 200


def test_name_clash_covers_inactive(client, admin_headers, flag_trailer):
    tid = flag_trailer["id"]
    client.put(f"/api/trailers/{tid}", headers=admin_headers, json={"is_active": False})
    r = client.post("/api/trailers", headers=admin_headers,
                    json={"name": flag_trailer["name"]})
    assert r.status_code == 400
    assert "Inactive" in r.json()["detail"]


def test_soft_deleted_hidden_even_with_include_inactive(client, admin_headers,
                                                        flag_trailer):
    tid = flag_trailer["id"]
    r = client.delete(f"/api/trailers/{tid}", headers=admin_headers)
    assert r.status_code == 200
    assert tid not in _trailer_ids(client, admin_headers)
    assert tid not in _trailer_ids(client, admin_headers, "?include_inactive=1")


def test_edit_reopen_backend_half_on_inactive_trailer(client, admin_headers, seeded):
    """The JS fallback needs: GET /api/calculations/{id} on a record whose trailer
    is inactive + GET /api/trailers/{tid} both to keep working."""
    tid = seeded["tt_id"]
    client.put(f"/api/trailers/{tid}", headers=admin_headers, json={"is_active": False})
    try:
        r = client.get(f"/api/calculations/{seeded['rec_id']}", headers=admin_headers)
        assert r.status_code == 200 and r.json()["trailer_type_id"] == tid
        r = client.get(f"/api/trailers/{tid}", headers=admin_headers)
        assert r.status_code == 200 and r.json()["is_active"] is False
        # BOM for an inactive trailer still loads (loadBOM path on edit reopen)
        r = client.get(f"/api/trailers/{tid}/bom", headers=admin_headers)
        assert r.status_code == 200
    finally:
        client.put(f"/api/trailers/{tid}", headers=admin_headers,
                   json={"is_active": True})
