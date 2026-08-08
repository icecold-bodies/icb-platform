"""3.2 m plywood+glue zero rule (scripts/rules/apply_32m_plywood_glue_zero.py).

collect() must: resolve a body by its name SET (exactly one active match);
pin exactly one 4MM PF PLYWOOD row per FRONT/SIDES section plus the GLUE LINE
row directly BELOW it (never the one above); fail loud — not guess — when the
row below the plywood is not a glue line; and be idempotent once guards are
written. The guard itself must zero the quantity at exactly 3.2 m and leave
every other length untouched, end-to-end through POST /api/calculate. The
export-side detector (app.services.zero_rule_note) must fire only when the
result's length matches a guarded row's pinned value.
"""
import uuid

import pytest


@pytest.fixture(scope="module")
def app_mod():
    import app.main as m
    from starlette.testclient import TestClient
    with TestClient(m.app) as _c:
        yield m


@pytest.fixture()
def staged(app_mod):
    """One synthetic body with the real FRONT/SIDES layup (ply flanked by two
    glue lines) plus a decoy body whose row below the plywood is NOT glue.

    Materials use the REAL names ('4MM PF PLYWOOD', 'GLUE LINE' — names are
    not unique in `materials`) so the script's name-matching is exercised
    verbatim; rows are tracked by id for teardown."""
    from app.database import BillOfMaterial, Material, SessionLocal, TrailerType
    mark = f"V32M-{uuid.uuid4().hex[:6]}"
    mat_ids, tt_ids = [], []
    with SessionLocal() as db:
        def mat(name):
            m = Material(name=name, unit_of_measure="m²", price_per_unit=10.0)
            db.add(m)
            db.flush()
            mat_ids.append(m.id)
            return m

        def row(tt, mat_row, section, sort, formula="(length+0.05)*(height+0.05)"):
            r = BillOfMaterial(trailer_type_id=tt.id, material_id=mat_row.id,
                               formula_expression=formula, waste_percentage=0,
                               bom_section=section, sort_order=sort)
            db.add(r)
            db.flush()
            return r

        t1 = TrailerType(name=f"{mark} MEDIUM CHILLER", is_active=True,
                         default_length=5.5, default_width=2.3, default_height=2.3)
        db.add(t1)
        db.flush()
        tt_ids.append(t1.id)
        ids = {"mark": mark, "t1": t1.id}
        for section, base in (("FRONT", 0), ("SIDES", 57)):
            key = section.lower()
            row(t1, mat(f"{mark} EXT SKIN"), section, base + 0)
            row(t1, mat("GLUE LINE"), section, base + 1)
            row(t1, mat(f"{mark} EPS"), section, base + 2)
            ids[f"{key}_glue_above"] = row(t1, mat("GLUE LINE"), section, base + 3).id
            ids[f"{key}_ply"] = row(t1, mat("4MM PF PLYWOOD"), section, base + 4).id
            ids[f"{key}_glue_below"] = row(t1, mat("GLUE LINE"), section, base + 5).id
            row(t1, mat(f"{mark} PICTURE FRAME"), section, base + 6)

        # Decoy: the row below the plywood is a frame, not glue → must fail loud.
        t2 = TrailerType(name=f"{mark} BAD LAYUP", is_active=True)
        db.add(t2)
        db.flush()
        tt_ids.append(t2.id)
        ids["t2"] = t2.id
        row(t2, mat("GLUE LINE"), "FRONT", 1)
        row(t2, mat("4MM PF PLYWOOD"), "FRONT", 2)
        row(t2, mat(f"{mark} PICTURE FRAME"), "FRONT", 3)

        db.commit()
    yield ids
    with SessionLocal() as db:
        from sqlalchemy import text
        db.execute(text("DELETE FROM icb_costings.bill_of_materials "
                        "WHERE trailer_type_id = ANY(:t)"), {"t": tt_ids})
        db.execute(text("DELETE FROM icb_costings.materials WHERE id = ANY(:m)"),
                   {"m": mat_ids})
        db.execute(text("DELETE FROM icb_costings.trailer_types WHERE id = ANY(:t)"),
                   {"t": tt_ids})
        db.commit()


def _rules_for(staged, key="t1", body="CHILLER MEDIUM"):
    from app.database import SessionLocal, TrailerType
    with SessionLocal() as db:
        name = db.get(TrailerType, staged[key]).name
    return [{"body": body, "names": {name}}]


def test_collect_pins_ply_and_glue_below_only(staged):
    from app.database import SessionLocal
    from scripts.rules.apply_32m_plywood_glue_zero import collect
    with SessionLocal() as db:
        actions, already, errors = collect(db, rules=_rules_for(staged))
        assert errors == []
        assert already == []
        got = {(a["section"], a["role"]): a["row"].id for a in actions}
        assert got == {
            ("FRONT", "plywood"): staged["front_ply"],
            ("FRONT", "glue"): staged["front_glue_below"],
            ("SIDES", "plywood"): staged["sides_ply"],
            ("SIDES", "glue"): staged["sides_glue_below"],
        }
        # The glue row ABOVE the plywood must never be pinned.
        pinned = {a["row"].id for a in actions}
        assert staged["front_glue_above"] not in pinned
        assert staged["sides_glue_above"] not in pinned
        for a in actions:
            assert a["after"] == f"({a['before']}) * (0 if abs(length - 3.2) < 1e-9 else 1)"


def test_bad_layup_fails_loud_and_writes_nothing(staged):
    from app.database import SessionLocal
    from scripts.rules.apply_32m_plywood_glue_zero import collect
    with SessionLocal() as db:
        actions, _already, errors = collect(
            db, rules=_rules_for(staged, key="t2", body="BAD LAYUP"))
        assert any("below the plywood" in e for e in errors)
        assert actions == []   # never guess a pair


def test_apply_is_idempotent_and_guard_zeroes_at_exactly_3_2(staged):
    from app.database import BillOfMaterial, SessionLocal
    from app.formula_engine import build_geometry, evaluate_formula
    from scripts.rules.apply_32m_plywood_glue_zero import collect
    with SessionLocal() as db:
        actions, _already, errors = collect(db, rules=_rules_for(staged))
        assert errors == [] and len(actions) == 4
        for a in actions:
            a["row"].formula_expression = a["after"]
        db.commit()
    with SessionLocal() as db:
        actions, already, errors = collect(db, rules=_rules_for(staged))
        assert errors == []
        assert actions == []                      # second run is a no-op
        assert {e["row"].id for e in already} == {
            staged["front_ply"], staged["front_glue_below"],
            staged["sides_ply"], staged["sides_glue_below"]}
        guarded = db.get(BillOfMaterial, staged["front_ply"]).formula_expression
    for length, expect_zero in ((3.2, True), (3.20, True), (3.19, False),
                                (3.21, False), (3.5, False)):
        err = []
        qty = evaluate_formula(guarded, build_geometry(
            {"length": length, "width": 2.3, "height": 2.3}), {}, {}, _err=err)
        assert not err, f"guard must not raise at L={length}"
        assert (qty == 0.0) is expect_zero, f"L={length}: qty={qty}"


def test_note_reaches_all_three_document_renderers():
    """build_doc_ctx carries zero_rule_note, and each of the three renderers
    (xlsx / docx / pdf) emits it — so the red line can't silently go missing
    from one format. Absent note ⇒ byte-identical-shaped document (no row)."""
    import io
    import zipfile

    import openpyxl

    from app.routers.exports import _render_docx, _render_pdf, _render_xlsx
    from app.services.document_context import build_doc_ctx

    note = "3.2 m rule applied: plywood + glue at R0,00 (FRONT, SIDES)"
    guard = "(1) * (0 if abs(length - 3.2) < 1e-9 else 1)"
    def item(cat, mat):
        return {"category": cat, "material": mat, "material_code": "",
                "formula": guard, "quantity": 0, "unit": "m²",
                "unit_price": 71.48, "waste_pct": 0, "line_cost": 0}

    result = {
        "items": [item("FRONT", "4MM PF PLYWOOD"), item("FRONT", "GLUE LINE"),
                  item("SIDES", "4MM PF PLYWOOD"), item("SIDES", "GLUE LINE")],
        "category_totals": {"FRONT": 0.0, "SIDES": 0.0}, "grand_total": 100.0,
        "geometry": {"length": 3.2},
    }
    common = dict(mode="preview", heading="Testing — X", sub="", client_name="",
                  spec_pairs=[("Length", "3.2 m")], spec_options=[],
                  ratios=[0.55], detail="items", generated_at="08 Aug 2026")
    ctx = build_doc_ctx(result=result, **common)
    assert ctx["zero_rule_note"] == note

    off = dict(result, geometry={"length": 3.5})
    ctx_off = build_doc_ctx(result=off, **common)
    assert ctx_off["zero_rule_note"] is None

    vals = [c.value for row in openpyxl.load_workbook(
        io.BytesIO(_render_xlsx(ctx).getvalue())).active.iter_rows() for c in row]
    assert note in vals
    vals_off = [c.value for row in openpyxl.load_workbook(
        io.BytesIO(_render_xlsx(ctx_off).getvalue())).active.iter_rows() for c in row]
    assert note not in vals_off

    def docx_xml(buf):
        with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as z:
            return z.read("word/document.xml").decode("utf-8", "ignore")
    assert note in docx_xml(_render_docx(ctx))
    assert note not in docx_xml(_render_docx(ctx_off))

    pdf_on, pdf_off = _render_pdf(ctx), _render_pdf(ctx_off)
    assert pdf_on[:4] == b"%PDF" and pdf_off[:4] == b"%PDF"
    assert len(pdf_on) > len(pdf_off)   # the extra red line is rendered


def test_api_calculate_zeroes_rows_visible_at_3_2(staged, app_mod):
    """End-to-end: guarded template through POST /api/calculate — the four
    rows stay IN the response with qty 0 / line 0 at 3.2 m and cost normally
    at 3.5 m."""
    import secrets

    from starlette.testclient import TestClient

    from app.database import SessionLocal, User, UserSession
    from scripts.rules.apply_32m_plywood_glue_zero import collect

    with SessionLocal() as db:
        actions, _already, errors = collect(db, rules=_rules_for(staged))
        assert errors == []
        for a in actions:                     # idempotent if the apply test ran first
            a["row"].formula_expression = a["after"]
        sid, csrf = str(uuid.uuid4()), secrets.token_hex(16)
        user = User(username=f"{staged['mark']}-user", password_hash="x", role="admin")
        db.add(user)
        db.flush()
        db.add(UserSession(id=sid, user_id=user.id, role=user.role,
                           csrf_token=csrf, expires_at=None))
        user_id = user.id
        db.commit()

    try:
        with TestClient(app_mod.app) as client:
            def calc(length):
                r = client.post(
                    "/api/calculate",
                    json={"trailer_type_id": staged["t1"],
                          "dimensions": {"length": length, "width": 2.3, "height": 2.3}},
                    headers={"Cookie": f"session_id={sid}", "X-CSRF-Token": csrf})
                assert r.status_code == 200, r.text
                return r.json()

            pinned = {staged["front_ply"], staged["front_glue_below"],
                      staged["sides_ply"], staged["sides_glue_below"]}

            at32 = calc(3.2)
            rows32 = {it["bom_id"]: it for it in at32["items"] if it["bom_id"] in pinned}
            assert set(rows32) == pinned          # rows VISIBLE, never hidden
            for it in rows32.values():
                assert it["quantity"] == 0 and it["line_cost"] == 0
                assert not it["formula_error"]

            at35 = calc(3.5)
            rows35 = {it["bom_id"]: it for it in at35["items"] if it["bom_id"] in pinned}
            assert set(rows35) == pinned
            for it in rows35.values():
                assert it["quantity"] > 0 and it["line_cost"] > 0
                assert not it["formula_error"]

            from app.services import zero_rule_note
            note32 = zero_rule_note(at32)
            assert note32 == "3.2 m rule applied: plywood + glue at R0,00 (FRONT, SIDES)"
            assert zero_rule_note(at35) is None
    finally:
        with SessionLocal() as db:
            from sqlalchemy import text
            db.execute(text("DELETE FROM icb_costings.user_sessions WHERE id = :s"),
                       {"s": sid})
            db.execute(text("DELETE FROM icb_costings.users WHERE id = :u"),
                       {"u": user_id})
            db.commit()
