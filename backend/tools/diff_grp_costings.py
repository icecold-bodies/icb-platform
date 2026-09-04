r"""
tools/diff_grp_costings.py
──────────────────────────
Independent row-level diff of two saved GRP Costings workbooks (old vs new),
plus reconciliation against a BA-produced change manifest.

Built for the September 2026 price update (v1.52): ratified default 1 says the
import may only run once TWO independently-derived diffs agree — this tool is
the second derivation. It reads cached values only (data_only=True) and never
writes to either workbook.

Layout discovery
────────────────
The workbook mixes three column layouts, sometimes inside one sheet:
  standard  desc=A  section=B  PRICE=G  TOTAL=H   (most body sheets)
  compact   desc=A  section=A  PRICE=E  TOTAL=F   (BAKERY BODIES, Sheet1,
                                                   TAUT LINER RIGID, AELER
                                                   PANELS, Manni DF)
  offset    desc=L  section=M  PRICE=R  TOTAL=S   (" 4.9 & UP CHILLER AND
                                                   2.5 WIDE " — cols A..H of
                                                   that sheet are #REF! junk,
                                                   including junk PRICE labels,
                                                   so the RIGHTMOST label on a
                                                   row wins)
Rather than hard-coding sheets to layouts, every row containing a PRICE label
re-anchors the columns for the rows beneath it. A sheet whose sections use
different label positions therefore parses correctly section by section.

Rows above a sheet's first label row are the dims / BODY OPTIONS / margin
header block, which the BA manifest deliberately excludes. Changes there are
collected separately (header_block_changes) and printed for the report — they
are real observations, just not part of the row-level import.

The manifest is the union of CHANGED rows and HIGHLIGHTED rows, so a row whose
only signal is Burt's orange fill (no value change) is emitted too.

Highlight detection
───────────────────
Burt marks new-since-last-issue rows with an orange fill (theme colour 9,
tint 0). A row counts as highlighted when its desc, price or total cell in
the NEW workbook carries that fill.

Usage
─────
    python tools/diff_grp_costings.py \
        --old  "...\August costings\GRP Costings 2018.xlsx" \
        --new  "...\September costings\GRP Costings 2018.xlsx" \
        --out  my_diff.csv \
        --reconcile docs/audit/september_price_update/september_change_manifest.csv

Exit code 0 = reconciliation clean (or no --reconcile); 1 = discrepancies.
"""
from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass

import openpyxl
from openpyxl.utils import get_column_letter

# Sheets that are planning/sales views, not body BOMs. Never diffed.
NON_BODY_SHEETS = {
    "TRAILER UNITS SOLD", "CHASSIS COSTINGS", "SHEET PLANNING",
    "VACUUM PLANNING HEIDELBERG", "SIDE TIPPER COSTINGS", "TRAILER PRICE LIST",
}

PRICE_TOLERANCE = 0.005   # rand — closer than this counts as unchanged
HIGHLIGHT_THEME = 9       # Burt's orange: theme colour 9 ...
HIGHLIGHT_TINT = 0.0      # ... at tint 0


@dataclass
class RowCells:
    desc: object
    price: object
    total: object
    section: str | None


def _num(v):
    """Cached cell value → float, or None when it isn't a number ('#REF!',
    labels, blanks). Error strings must not read as zero. Some price cells are
    TEXT-formatted numbers ('116.1500') — those must still read as numbers."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except ValueError:
            return None
    return None


def _txt(v):
    return v.strip() if isinstance(v, str) else None


def _is_highlighted(cell) -> bool:
    try:
        f = cell.fill
        if f is None or f.patternType is None:
            return False
        c = f.start_color
        if c is None or getattr(c, "type", None) != "theme":
            return False
        return c.theme == HIGHLIGHT_THEME and float(c.tint or 0) == HIGHLIGHT_TINT
    except Exception:
        return False


def _price_label_col(row) -> int | None:
    """Column of the PRICE label on this row, or None. Takes the RIGHTMOST
    match: the offset CHILLER sheet's dead A..H region still contains the old
    header labels, and anchoring to those would walk the #REF! junk."""
    found = None
    for cell in row:
        if isinstance(cell.value, str) and cell.value.strip().upper() == "PRICE":
            found = cell.column
    return found


def parse_sheet(ws) -> tuple[dict[int, RowCells], int | None]:
    """Walk one sheet; return ({row_number: RowCells}, first_label_row) for
    every row that has a text in the current desc column. Label rows re-anchor
    the layout."""
    # Sheet-level default: standard layout.
    desc_col, price_col, total_col = 1, 7, 8
    section = None
    out: dict[int, RowCells] = {}
    first_label_row: int | None = None

    max_col = min(ws.max_column or 30, 40)
    for row in ws.iter_rows(min_row=1, max_col=max_col):
        rn = row[0].row

        # Label row? A cell reading PRICE re-anchors columns for rows below.
        price_label_col = _price_label_col(row)
        if price_label_col is not None:
            if first_label_row is None:
                first_label_row = rn
            price_col = price_label_col
            total_col = price_label_col + 1
            desc_col = 12 if price_label_col >= 12 else 1
            # Section name: leftmost string on the label row left of the labels
            # (compact/standard), else one row up (offset layout).
            section = None
            for cell in row:
                if cell.column >= price_label_col:
                    break
                v = _txt(cell.value)
                if v and v.upper() not in {"WIDTH", "HEIGHT", "M2", "QTY",
                                           "QUANTITY", "QUAN", "LENGTH",
                                           "DESCRIPTION", "MATERIAL", "ITEM"}:
                    section = v
                    break
            if section is None and rn > 1:
                above = ws.cell(rn - 1, desc_col + 1).value or ws.cell(rn - 1, desc_col).value
                section = _txt(above)
            continue

        desc_cell = row[desc_col - 1] if len(row) >= desc_col else None
        desc = _txt(desc_cell.value) if desc_cell is not None else None
        if not desc:
            continue
        price_cell = row[price_col - 1] if len(row) >= price_col else None
        total_cell = row[total_col - 1] if len(row) >= total_col else None
        out[rn] = RowCells(
            desc=desc,
            price=price_cell.value if price_cell is not None else None,
            total=total_cell.value if total_cell is not None else None,
            section=section,
        )
    return out, first_label_row


def parse_highlights(ws) -> set[int]:
    """Row numbers whose desc/price/total cell carries Burt's orange fill.
    Runs the same layout walk so the columns match parse_sheet."""
    desc_col, price_col, total_col = 1, 7, 8
    out: set[int] = set()
    max_col = min(ws.max_column or 30, 40)
    for row in ws.iter_rows(min_row=1, max_col=max_col):
        rn = row[0].row
        price_label_col = _price_label_col(row)
        if price_label_col is not None:
            price_col = price_label_col
            total_col = price_label_col + 1
            desc_col = 12 if price_label_col >= 12 else 1
            continue
        cells = []
        for col in (desc_col, price_col, total_col):
            if len(row) >= col:
                cells.append(row[col - 1])
        if any(_is_highlighted(c) for c in cells):
            out.add(rn)
    return out


def diff_workbooks(old_path: str, new_path: str) -> tuple[list[dict], list[dict]]:
    """Returns (bom_zone_changes, header_block_changes)."""
    wb_old = openpyxl.load_workbook(old_path, data_only=True)
    wb_new = openpyxl.load_workbook(new_path, data_only=True)
    changes: list[dict] = []
    header_changes: list[dict] = []

    for sheet in wb_new.sheetnames:
        if sheet in NON_BODY_SHEETS:
            continue
        if sheet not in wb_old.sheetnames:
            changes.append({"sheet": sheet, "row": 0, "note": "SHEET ONLY IN NEW"})
            continue
        old_rows, first_old = parse_sheet(wb_old[sheet])
        new_rows, first_new = parse_sheet(wb_new[sheet])
        highlights = parse_highlights(wb_new[sheet])
        # BOM zone starts at the first label row; anything above is the
        # dims / BODY OPTIONS / margin header block the manifest excludes.
        bom_start = min(x for x in (first_old, first_new, 10**9) if x is not None)

        for rn in sorted(set(old_rows) | set(new_rows)):
            o, n = old_rows.get(rn), new_rows.get(rn)
            desc_old = o.desc if o else None
            desc_new = n.desc if n else None
            po, pn = _num(o.price) if o else None, _num(n.price) if n else None
            to, tn = _num(o.total) if o else None, _num(n.total) if n else None

            desc_changed = (desc_old or "") != (desc_new or "")
            price_changed = _num_changed(po, pn, o.price if o else None, n.price if n else None)
            total_changed = _num_changed(to, tn, o.total if o else None, n.total if n else None)
            any_change = desc_changed or price_changed or total_changed
            if not any_change and rn not in highlights:
                continue
            rec = {
                "sheet": sheet,
                "section": (n.section if n else None) or (o.section if o else None) or "",
                "row": rn,
                "desc_old": desc_old or "",
                "desc_new": desc_new or "",
                "price_old": "" if po is None else repr(po),
                "price_new": "" if pn is None else repr(pn),
                "total_old": "" if to is None else repr(to),
                "total_new": "" if tn is None else repr(tn),
                "desc_changed": desc_changed,
                "price_changed": price_changed,
                "total_changed": total_changed,
                "highlighted": rn in highlights,
            }
            if rn < bom_start:
                if any_change:
                    header_changes.append(rec)
                continue
            changes.append(rec)
    wb_old.close()
    wb_new.close()
    return changes, header_changes


def _num_changed(a: float | None, b: float | None, raw_a, raw_b) -> bool:
    """Numeric change with tolerance. A #REF!/text on either side only counts
    as a change when the raw strings differ (e.g. number → #REF!)."""
    if a is not None and b is not None:
        return abs(a - b) > PRICE_TOLERANCE
    if a is None and b is None:
        ra = raw_a.strip() if isinstance(raw_a, str) else raw_a
        rb = raw_b.strip() if isinstance(raw_b, str) else raw_b
        return ra != rb
    return True


# ── Reconciliation against the BA manifest ──────────────────────────────────

def _man_num(s: str):
    s = (s or "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def reconcile(my_rows: list[dict], manifest_path: str) -> list[str]:
    """Compare my diff against the BA manifest keyed on (sheet, row).
    Returns a list of discrepancy strings — empty means the two agree."""
    manifest = {}
    with open(manifest_path, encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            manifest[(r["sheet"], int(r["row"]))] = r

    mine = {(r["sheet"], r["row"]): r for r in my_rows if r.get("row")}

    problems: list[str] = []
    for key in sorted(set(manifest) | set(mine)):
        m, d = manifest.get(key), mine.get(key)
        if m is None:
            problems.append(f"ONLY IN MY DIFF   {key}: {d['desc_old']!r}->{d['desc_new']!r} "
                            f"p {d['price_old']}->{d['price_new']} t {d['total_old']}->{d['total_new']}")
            continue
        if d is None:
            problems.append(f"ONLY IN MANIFEST  {key}: {m['desc_old']!r}->{m['desc_new']!r} "
                            f"p {m['price_old']}->{m['price_new']}")
            continue
        for flag in ("desc_changed", "price_changed", "total_changed", "highlighted"):
            if str(d[flag]) != m[flag]:
                problems.append(f"FLAG MISMATCH     {key} {flag}: mine={d[flag]} manifest={m[flag]}")
        # The manifest keeps descs verbatim (incl. trailing spaces); this tool
        # strips them. Compare stripped — the raw form still lives in each CSV.
        if str(d["desc_old"]).strip() != m["desc_old"].strip() \
                or str(d["desc_new"]).strip() != m["desc_new"].strip():
            problems.append(f"DESC MISMATCH     {key}: mine {d['desc_old']!r}->{d['desc_new']!r} "
                            f"manifest {m['desc_old']!r}->{m['desc_new']!r}")
        for col in ("price_old", "price_new", "total_old", "total_new"):
            mv, dv = _man_num(m[col]), _man_num(d[col])
            if (mv is None) != (dv is None) or (mv is not None and abs(mv - dv) > PRICE_TOLERANCE):
                problems.append(f"VALUE MISMATCH    {key} {col}: mine={d[col]} manifest={m[col]}")
    return problems


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[2])
    ap.add_argument("--old", required=True, help="baseline workbook (August)")
    ap.add_argument("--new", required=True, help="updated workbook (September)")
    ap.add_argument("--out", help="write my diff rows to this CSV")
    ap.add_argument("--reconcile", help="BA manifest CSV to reconcile against")
    ap.add_argument("--prove", action="store_true",
                    help="self-test: fail unless the diff contains the known "
                         "'Adv Vacuum panels' row 35 change (359 -> 409)")
    args = ap.parse_args()

    rows, header_changes = diff_workbooks(args.old, args.new)
    real = [r for r in rows if r.get("row")]
    print(f"my diff: {len(real)} changed/highlighted rows "
          f"(price={sum(1 for r in real if r['price_changed'])}, "
          f"desc={sum(1 for r in real if r['desc_changed'])}, "
          f"total={sum(1 for r in real if r['total_changed'])}, "
          f"highlighted={sum(1 for r in real if r['highlighted'])})")

    if header_changes:
        print(f"\nheader-block changes OUTSIDE the manifest's row zone "
              f"({len(header_changes)} rows — dims/options/margin block; "
              f"report-only, never imported):")
        for r in header_changes:
            if r["desc_changed"] or r["price_changed"]:
                print(f"  {r['sheet']!r} row {r['row']}: {r['desc_old']!r}->{r['desc_new']!r} "
                      f"price {r['price_old']}->{r['price_new']}")
        n_total_only = sum(1 for r in header_changes
                           if not (r["desc_changed"] or r["price_changed"]))
        print(f"  (+{n_total_only} header rows where only the running total moved)")

    if args.prove:
        hits = [r for r in real if r["sheet"] == "Adv Vacuum panels" and r["row"] == 35]
        ok = bool(hits) and _man_num(hits[0]["price_old"]) == 359.0 \
            and _man_num(hits[0]["price_new"]) == 409.0
        print(f"prove-known-change (Adv Vacuum panels row 35, 359->409): "
              f"{'DETECTED' if ok else '*** NOT DETECTED ***'}")
        if not ok:
            return 1

    if args.out:
        with open(args.out, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=[
                "sheet", "section", "row", "desc_old", "desc_new",
                "price_old", "price_new", "total_old", "total_new",
                "desc_changed", "price_changed", "total_changed", "highlighted"])
            w.writeheader()
            for r in real:
                w.writerow({k: r.get(k, "") for k in w.fieldnames})
        print(f"wrote {args.out}")

    if args.reconcile:
        problems = reconcile(real, args.reconcile)
        if problems:
            print(f"\nRECONCILIATION: {len(problems)} discrepancies")
            for p in problems:
                print("  " + p)
            return 1
        print("\nRECONCILIATION: clean — my diff and the manifest agree.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
